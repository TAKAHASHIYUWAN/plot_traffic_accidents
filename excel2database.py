import openpyxl as excel
import sqlalchemy as sa
import os 
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

database_file = os.path.join(os.path.abspath(os.getcwd()) , 'honhyo.db')

WIN_DESKTOP_PATH = "/mnt/c/Users/yuwan/Desktop/"

# windowsのexcelデータを読み込む。
# honhyoは「本票」を意味する。
wb_honhyo = excel.load_workbook(filename = WIN_DESKTOP_PATH + "honhyo.xlsx")
ws_honhyo = wb_honhyo.active
maxRow = ws_honhyo.max_row
maxClm = ws_honhyo.max_column



#データベースを作成
conn = sa.create_engine('sqlite:///honhyo.db',)

Base = declarative_base()
class Honhyo(Base):
    __tablename__ = "honhyo"
    #excelからカラムの名前を抽出

    id = sa.Column("id",sa.Integer, primary_key = True)
    prefecture = sa.Column("prefecture",sa.String )
    incident = sa.Column('incident', sa.Integer )
    dead = sa.Column('dead', sa.Integer  )
    injured = sa.Column('injured', sa.Integer  )
    year = sa.Column('year', sa.Integer  )
    month = sa.Column('month', sa.Integer  )
    day = sa.Column('day', sa.Integer  )
    hours = sa.Column('hours', sa.Integer  )
    minutes = sa.Column('minutes', sa.Integer  )
    latitude = sa.Column('latitude', sa.Float  )
    longitude = sa.Column('longitude', sa.Float  )
    
    def __init__(self,prefecture,incident,dead,injured,year,month,day,hours,minutes,latitude,longitude):
        self.prefecture = prefecture
        self.incident = incident
        self.dead = dead
        self.injured = injured
        self.year = year
        self.month = month
        self.day = day
        self.hours = hours
        self.minutes = minutes
        self.latitude = latitude
        self.longitude = longitude

    def __repr__(self):  
            return "<Honhyo({},{},{},{},{},{},{},{},{},{})>".format(self.prefecture,self.incident,
                                        self.dead,self.injured,self.year,self.month,self.day,
                                        self.hours,self.minutes,self.latitude,self.longitude)

Base.metadata.create_all(conn)


"""
DBにデータを格納するためには、
first_instance = Honhyo(prefecture1,incident1,......)
second_instance = Honhyo(prefecture2,incident2,......)
......


session.add(first_instance),session_add(second_instance)か
session.add_all( [first_instance,second_instance,.....] )
のようにする必要がある。

instance_list = [firsr_instance, second_instance,........]
"""
instance_list = []
for rows in range(2, maxRow + 1 ):
    cell_list = []
    for clms in range(1,maxClm + 1 ):
        cellValue = ws_honhyo.cell(row=rows,column=clms).value
        cell_list.append(cellValue)
    instance_list.append(Honhyo(cell_list[0],cell_list[1],cell_list[2],
                        cell_list[3],cell_list[4],cell_list[5],cell_list[6],
                        cell_list[7],cell_list[8],cell_list[9],cell_list[10]))



Session = sessionmaker(bind = conn)
session = Session()
session.add_all(instance_list)

session.commit()
